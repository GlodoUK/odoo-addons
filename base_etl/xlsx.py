"""XLSX codec for ETL steps: a file handle <-> a list of row dicts.

The Excel mirror of :mod:`.csv`, reading and writing one worksheet with the
first row as the header. Both functions take an open **binary** file handle
(an xlsx is a zip archive -- there is no text mode); ``open(..., "rb"/"wb")``,
``fsspec.open(..., "rb"/"wb")`` and ``io.BytesIO()`` all work, and the caller
owns opening/closing. openpyxl is a hard dependency of base_etl (see the
manifest). Unlike the CSV codec's all-strings, values round-trip as whatever
types Excel carries (str/int/float/datetime). Nothing here imports Odoo, so it
is unit-testable on an in-memory handle.
"""

import openpyxl


def read_rows(handle, *, sheet=None):
    """Parse the workbook in ``handle`` into a list of dicts keyed by the
    header row of ``sheet`` (its title; the active sheet when omitted).

    ``handle`` is a readable binary file object. Read in read-only, data-only
    mode, so a large workbook streams and formula cells yield their last cached
    value rather than the formula text. An empty sheet (no header row) yields
    ``[]``. The rows are fully materialised before returning, so the caller's
    handle can close straight after::

        with fs.open(path, "rb") as handle:
            rows = read_rows(handle)
    """
    workbook = openpyxl.load_workbook(handle, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet] if sheet else workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return []
        return [dict(zip(header, row, strict=False)) for row in rows]
    finally:
        # Closes the workbook (and its zip), not the caller's handle.
        workbook.close()


def write_rows(handle, rows, *, fieldnames=None, sheet=None):
    """Write ``rows`` (an iterable of dicts) to ``handle`` as an xlsx workbook
    with a header row.

    ``handle`` is a writable binary file object. ``fieldnames`` defaults to the
    keys of the first row, in order; pass it explicitly to fix the column
    order/subset, or to emit just a header when ``rows`` is empty. ``sheet``
    names the worksheet. Written in write-only mode so large sets stream::

        with fs.open(path, "wb") as handle:
            write_rows(handle, rows)
    """
    rows = list(rows)
    if fieldnames is None:
        fieldnames = list(rows[0].keys()) if rows else []
    workbook = openpyxl.Workbook(write_only=True)
    worksheet = workbook.create_sheet(title=sheet)
    worksheet.append(list(fieldnames))
    for row in rows:
        worksheet.append([row.get(name) for name in fieldnames])
    workbook.save(handle)
